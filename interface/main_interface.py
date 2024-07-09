import sys
from PyQt6.QtWidgets import QApplication, QWidget, QMessageBox, QFileDialog
from PyQt6.QtWidgets import QTableWidgetItem
from Gerenciamento_interface import Ui_Form
from PyQt6.QtCore import QDate
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import sqlite3
import random
import datetime
from datetime import datetime


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)
        
        # Conectar os botões aos métodos correspondentes
        self.ui.Inicio.clicked.connect(self.mostrar_frame_inicio)
        self.ui.pushButton_18.clicked.connect(self.exportar_excel)
        self.ui.pushButton_24.clicked.connect(self.pesquisar_estoque)
        self.ui.produtos.clicked.connect(self.mostrar_frame_produto)
        self.ui.produtos.clicked.connect(self.carregar_dados_produto)
        self.ui.pushButton.clicked.connect(self.consultar_dados_produto)
        self.ui.entreda_2.clicked.connect(self.mostrar_frame_entrada)
        self.ui.cadastro_2.clicked.connect(self.mostrar_frame_cadastro)
        self.ui.saida_2.clicked.connect(self.mostrar_frame_saida)
        self.ui.relatorio.clicked.connect(self.mostrar_frame_relatorio)
        self.ui.relatorio.clicked.connect(self.relatorio_estoque)
        self.ui.suporte_2.clicked.connect(self.mostrar_frame_suporte)
        self.ui.pushButton_13.clicked.connect(self.gerar_certificado_os)
        self.ui.entreda_2.clicked.connect(self.atualizar_registro_produto)
        self.ui.salvar_registro_entrada.clicked.connect(self.atualizar_banco_produto)
        self.ui.pushButton_19.clicked.connect(self.pesquisar_produto_entrada)
        self.ui.pushButton_15.clicked.connect(self.excluir_registro_produto)
        self.ui.pushButton_11.clicked.connect(self.redirecionar_para_suporte)
        self.ui.pushButton_10.clicked.connect(self.redirecionar_para_configuracao)
        self.ui.pushButton_20.clicked.connect(self.importar_dados)
        # Conectar os botões aos métodos correspondentes
        self.ui.toolButton_3.clicked.connect(self.pesquisar_produto_por_codigo)
        self.ui.pushButton_2.clicked.connect(self.processar_saida)
        self.ui.pushButton_2.clicked.connect(self.cadastrar_saida)
        self.ui.pushButton_5.clicked.connect(self.excluir_item_selecionado_venda)
        self.ui.pushButton_6.clicked.connect(self.aplicar_desconto)
        self.ui.cadastra_cliente.clicked.connect(self.cadastrar_cliente)
        self.ui.castrar_fornecedor.clicked.connect(self.cadastrar_fornecedor)
        self.ui.pushButton_3.clicked.connect(self.adicionar_item_venda)
        self.ui.pushButton_4.clicked.connect(self.registrar_venda)
        # Conectar os botões aos slots
        self.ui.relatorio.clicked.connect(self.exibir_registros_venda)
        self.ui.pushButton_16.clicked.connect(self.aplicar_filtro_vendas)
        self.ui.pushButton_12.clicked.connect(self.exportar_vendas_excel)

        # Definir a data final para a data atual
        self.ui.dateEdit.setDate(QDate.currentDate())
         # Inicializa os valores dos labels
        self.ui.label_24.setText("Total: R$ 0.00")  # Subtotal
        self.ui.label_25.setText("0.00")            # Desconto
        self.ui.label_28.setText("Total: R$ 0.00")  # Total a Pagar
        # Conectar o botão 'Registrar Produto' ao método correspondente
        self.ui.botao_entreda_salvar.clicked.connect(self.cadastrar_produto)
        # Conexão com o banco de dados SQLite
        self.db_connection = sqlite3.connect("database/data_sys.db")
        self.cursor = self.db_connection.cursor()
        # Carregar fornecedores no QComboBox ao iniciar a aplicação
        self.carregar_fornecedores()
        self.carregar_clientes()
        # Função para redirecionar para a página de suporte
    def redirecionar_para_suporte(self):
        self.ui.pages_suporte.setCurrentWidget(self.ui.page)
    def redirecionar_para_configuracao(self):
        self.ui.pages_suporte.setCurrentWidget(self.ui.page_5)
    def mostrar_frame_inicio(self):
        self.ui.stackedWidget.setCurrentIndex(3)
    def mostrar_frame_produto(self):
        self.ui.stackedWidget.setCurrentIndex(1)
    def mostrar_frame_entrada(self):
        self.ui.stackedWidget.setCurrentIndex(0)
        self.carregar_fornecedores()  # Certifique-se de carregar os fornecedores ao abrir o frame de entrada
    def mostrar_frame_cadastro(self):
        self.ui.stackedWidget.setCurrentIndex(2)
    def mostrar_frame_saida(self):
        self.ui.stackedWidget.setCurrentIndex(4)
        numero_aleatorio = gerar_numero_aleatorio(1000, 9999)
        numero_aleatorioSR = str(numero_aleatorio)
        self.cod_venda = numero_aleatorioSR
        self.ui.label_53.setText(numero_aleatorioSR)
    def mostrar_frame_relatorio(self):
        self.ui.stackedWidget.setCurrentIndex(6)
    def mostrar_frame_suporte(self):
        self.ui.stackedWidget.setCurrentIndex(5)

    def cadastrar_cliente(self):
        # Coletando os dados do formulário
        codigo = self.ui.codigo_cliente.text()
        nome = self.ui.nome_cliente.text()
        telefone = self.ui.telefone_cliente.text()
        email = self.ui.email_cliente.text()
        cnpj_cpf = self.ui.cnpcpf_cliente.text()
        pessoa = self.ui.pessoa_cliente.currentText()
        endereco = self.ui.endereco_cliente.text()
        cep = self.ui.cep_cliente.text()
        cidade = self.ui.cidade_cliente.text()
        numero = self.ui.numero_cliente.text()
        complemento = self.ui.complemento_cliente.text()

        # Verificando se todos os campos foram preenchidos
        if not (codigo and nome and telefone and cnpj_cpf and pessoa and endereco and cep and cidade and numero):
            QMessageBox.warning(self, "Atenção", "Todos os campos devem ser preenchidos.")
            return

        # Inserindo os dados na tabela registro_cliente
        self.cursor.execute("""
            INSERT INTO registro_cliente 
            (codigo, nome, telefone, email, cnpj_cpf, pessoa, endereco, cep, cidade, numero, complemento) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (codigo, nome, telefone, email, cnpj_cpf, pessoa, endereco, cep, cidade, numero, complemento))

        # Commit das alterações
        self.db_connection.commit()

        # Exibir mensagem de sucesso
        QMessageBox.information(self, "Sucesso", "Cadastro de cliente realizado com sucesso.")

        # Limpar os campos após a inserção
        self.limpar_campos()

    def relatorio_estoque(self):
        # Consulta para obter todos os produtos do banco de dados
        self.cursor.execute("""
            SELECT id, codigo, nome, marca, data_entrada, grupo, peso, medida, fornecedor, lote, estoque, alerta_reposicao, valor_venda, valor_compra, descricao
            FROM registro_novo_produto
        """)
        produtos = self.cursor.fetchall()

        # Limpar a tabela antes de preencher
        self.ui.tableWidget_3.setRowCount(0)

        # Preencher a tabela com os produtos
        for row_number, produto in enumerate(produtos):
            self.ui.tableWidget_3.insertRow(row_number)
            for column_number, data in enumerate(produto):
                self.ui.tableWidget_3.setItem(row_number, column_number, QTableWidgetItem(str(data)))
    def pesquisar_estoque(self):
        # Obter o texto de pesquisa do lineEdit_15
        texto_pesquisa = self.ui.lineEdit_15.text().strip()

        # Verificar se o campo de pesquisa está vazio
        if not texto_pesquisa:
            QMessageBox.warning(self, "Erro", "Por favor, insira um texto para pesquisar.")
            return

        # Consulta para obter produtos que correspondam ao texto de pesquisa
        query = """
            SELECT id, codigo, nome, marca, data_entrada, grupo, peso, medida, fornecedor, lote, estoque, alerta_reposicao, valor_venda, valor_compra, descricao
            FROM registro_novo_produto
            WHERE nome LIKE ? OR codigo LIKE ? OR marca LIKE ? OR grupo LIKE ? OR fornecedor LIKE ?
        """
        # Adiciona o texto de pesquisa com coringas (%) para busca parcial
        parametros = (f'%{texto_pesquisa}%', f'%{texto_pesquisa}%', f'%{texto_pesquisa}%', f'%{texto_pesquisa}%', f'%{texto_pesquisa}%')
        self.cursor.execute(query, parametros)
        produtos = self.cursor.fetchall()

        # Verificar se a pesquisa retornou resultados
        if not produtos:
            QMessageBox.information(self, "Resultado", "Nenhum produto encontrado para a pesquisa realizada.")
            return

        # Limpar a tabela antes de preencher
        self.ui.tableWidget_3.setRowCount(0)

        # Preencher a tabela com os produtos filtrados
        for row_number, produto in enumerate(produtos):
            self.ui.tableWidget_3.insertRow(row_number)
            for column_number, data in enumerate(produto):
                self.ui.tableWidget_3.setItem(row_number, column_number, QTableWidgetItem(str(data)))


    def cadastrar_fornecedor(self):
        # Coletando os dados do formulário
        codigo = self.ui.edit_cod_forn.text()
        nome = self.ui.edit_forn_nome.text()
        telefone = self.ui.edit_forn_tele.text()
        email = self.ui.edit_email_.text()
        cnpj_cpf = self.ui.edit_forn_cnpjcpf.text()
        pessoa = self.ui.pessoa_forn.currentText()
        produto_servico = self.ui.prod_edit_forn.text()
        valor = self.ui.valor_edit_forn.text()
        mais_informacao = self.ui.mais_edit_forn.toPlainText()
        endereco = self.ui.ender_edit_forn.text()
        cep = self.ui.cep_edit_forn.text()
        cidade = self.ui.cid_edit_forn.text()
        numero = self.ui.num_edit_forn.text()
        complemento = self.ui.comp_edit_forn.text()

        # Verificando se todos os campos foram preenchidos
        if not (codigo and nome and telefone and cnpj_cpf and pessoa and produto_servico and valor and mais_informacao and endereco and cep and cidade and numero):
            QMessageBox.warning(self, "Atenção", "Todos os campos devem ser preenchidos.")
            return

        # Inserindo os dados na tabela registro_fornecedor
        self.cursor.execute("""
            INSERT INTO registro_fornecedor  
            (codigo, nome, telefone, email, cnpj_cpf, pessoa, produto_servico, valor, mais_informacao, endereco, cep, cidade, numero, complemento) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (codigo, nome, telefone, email, cnpj_cpf, pessoa, produto_servico, valor, mais_informacao, endereco, cep, cidade, numero, complemento))

        # Commit das alterações
        self.db_connection.commit()

        # Exibir mensagem de sucesso
        QMessageBox.information(self, "Sucesso", "Cadastro de fornecedor realizado com sucesso.")

        # Limpar os campos após a inserção
        self.limpar_campos()
    def pesquisar_produto_por_codigo(self):
        # Obter o código do produto do lineEdit_7
        codigo_produto = self.ui.lineEdit_7.text()
        
        # Pesquisar no banco de dados
        self.cursor.execute("SELECT * FROM registro_novo_produto WHERE codigo = ?", (codigo_produto,))
        produto = self.cursor.fetchone()
        
        # Verificar se o produto foi encontrado
        if produto:
            # Preencher os lineEdits com as informações do produto encontrado
            self.ui.lineEdit_10.setText(produto[2])  # nome
            self.ui.lineEdit_3.setText(produto[5])   # grupo
            self.ui.lineEdit_4.setText(produto[14])
            self.ui.volume_atual_saida.setText(str(produto[10]))  # estoque
            self.ui.volume_minimo_saida.setText(str(produto[11]))  # alerta_reposicao
            self.ui.lote_saida.setText(produto[9])  # lote
            self.ui.peso_saida.setText(str(produto[6]))  # peso
            self.ui.medida_saida.setText(str(produto[7]))  # medida
        else:
            QMessageBox.warning(self, "Atenção", "Código não encontrado.")
            # Se o produto não for encontrado, limpar os lineEdits
            self.ui.lineEdit_10.clear()
            self.ui.lineEdit_3.clear()
            self.ui.volume_atual_saida.clear()
            self.ui.volume_minimo_saida.clear()
            self.ui.lote_saida.clear()
            self.ui.peso_saida.clear()
            self.ui.medida_saida.clear()
    def processar_saida(self):
        # Obter a quantidade de saída do lineEdit_8
        quantidade_saida_str = self.ui.lineEdit_8.text()

        # Verificar se a quantidade de saída é válida
        try:
            quantidade_saida = int(quantidade_saida_str)
        except ValueError:
            QMessageBox.warning(self, "Atenção", "Quantidade inválida.")
            return

        # Obter o estoque atual do banco de dados
        codigo_produto = self.ui.lineEdit_7.text()
        self.cursor.execute("SELECT estoque FROM registro_novo_produto WHERE codigo = ?", (codigo_produto,))
        estoque_atual_str = self.cursor.fetchone()[0]

        try:
            # Convertendo estoque_atual para inteiro
            estoque_atual = int(estoque_atual_str)
        except ValueError:
            QMessageBox.warning(self, "Atenção", "Erro ao obter o estoque atual.")
            return

        # Verificar se há estoque suficiente
        if estoque_atual >= quantidade_saida:
            # Subtrair a quantidade de saída do estoque atual
            estoque_atual -= quantidade_saida

            # Atualizar o estoque no banco de dados
            self.cursor.execute("UPDATE registro_novo_produto SET estoque = ? WHERE codigo = ?", (estoque_atual, codigo_produto))
            self.db_connection.commit()

            QMessageBox.information(self, "Sucesso", f"{quantidade_saida} volumes retirados do estoque.")
            
        else:
            QMessageBox.warning(self, "Atenção", "Estoque insuficiente.")
    def cadastrar_saida(self):
        # Obter os valores dos campos de entrada
        data_saida = datetime.now().strftime("%Y-%m-%d")  
        codigo = self.ui.lineEdit_7.text()
        produto = self.ui.lineEdit_10.text()
        grupo = self.ui.lineEdit_3.text()
        medida = self.ui.medida_saida.text()
        peso = self.ui.peso_saida.text()
        volume_expedido = self.ui.lineEdit_8.text()
        volume_atual = self.ui.volume_atual_saida.text()
        destino = self.ui.lineEdit_5.text()  # Certifique-se de ter um campo para o destino da saída
        descricao = self.ui.lineEdit_4.text()

        # Verificar se todos os campos obrigatórios foram preenchidos
        if not (codigo and produto and grupo and volume_expedido):
            QMessageBox.warning(self, "Atenção", "Todos os campos (*) devem ser preenchidos.")
            return

        # Inserir os dados na tabela registro_saida
        self.cursor.execute("""
            INSERT INTO registro_saida 
            (codigo, produto, grupo, medida, data_saida, peso, volume_expedido, volume_atual, destino, descricao) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (codigo, produto, grupo, medida, data_saida, peso, volume_expedido, volume_atual, destino, descricao))

        # Commit das alterações
        self.db_connection.commit()

        # Exibir mensagem de sucesso
        QMessageBox.information(self, "Sucesso", "Saída cadastrada com sucesso.")
        self.limpar_campos()

    def cadastrar_produto(self):
        # Coletando os dados do formulário
        codigo = self.ui.codigo.text()
        nome = self.ui.nome.text()
        marca = self.ui.marca.text()
        grupo = self.ui.grupo.text()
        peso = self.ui.peso.text()
        medida = self.ui.medida.text()
        fornecedor = self.ui.fornecedor.currentText()
        lote = self.ui.lote_produto.text()
        estoque = self.ui.estoque.text()
        data_entrada = datetime.now().strftime("%Y-%m-%d")
        alerta_reposicao = self.ui.reposicao.text()
        valor_venda = self.ui.venda.text()
        valor_compra = self.ui.lineEdit_2.text()
        descricao = self.ui.descricao.text()

        # Verificando se todos os campos foram preenchidos
        if not (codigo and nome and marca and grupo and peso and medida and fornecedor and lote and estoque and alerta_reposicao and valor_venda and valor_compra and descricao):
            QMessageBox.warning(self, "Atenção", "Todos os campos devem ser preenchidos.")
            return

        # Inserindo os dados na tabela registro_novo_produto
        self.cursor.execute("""
                    INSERT INTO registro_novo_produto 
                    (codigo, nome, marca, data_entrada, grupo, peso, medida, fornecedor, lote, estoque,alerta_reposicao , valor_venda, valor_compra, descricao)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (codigo, nome, marca, data_entrada, grupo, peso, medida, fornecedor, lote, estoque, alerta_reposicao, valor_venda, valor_compra, descricao))
        # Commit das alterações
        self.db_connection.commit()
                # Limpar os campos após a inserção
        self.limpar_campos()
    
        # Exibir mensagem de sucesso
        QMessageBox.information(self, "Sucesso", "Cadastro de produto realizado com sucesso.")

    def carregar_fornecedores(self):
        self.cursor.execute("SELECT nome FROM registro_fornecedor")
        fornecedores = self.cursor.fetchall()
        
        self.ui.fornecedor.clear()
        for fornecedor in fornecedores:
            self.ui.fornecedor.addItem(fornecedor[0])
    def carregar_clientes(self):
        self.cursor.execute("SELECT nome FROM registro_cliente")
        clientes = self.cursor.fetchall()
        
        self.ui.comboBox_3.clear()
        for cliente in clientes:
            self.ui.comboBox_3.addItem(cliente[0])
    def adicionar_item_venda(self):
        codigo_produto = self.ui.lineEdit_6.text()  # Mudança para lineEdit_7
        quantidade_expedida = self.ui.lineEdit_9.text()
        
        # Verifica se a quantidade expedida é válida
        if not quantidade_expedida.isdigit():
            QMessageBox.warning(self, "Erro", "Quantidade expedida inválida.")
            return
        quantidade_expedida = int(quantidade_expedida)  # Converte para inteiro após a verificação

        # Verifica se o código do produto está preenchido
        if not codigo_produto:
            QMessageBox.warning(self, "Erro", "Por favor, insira o código do produto.")
            return

        # Consulta o banco de dados para verificar se o produto existe
        self.cursor.execute("SELECT nome, estoque, valor_venda, descricao FROM registro_novo_produto WHERE codigo = ?", (codigo_produto,))
        produto = self.cursor.fetchone()

        if not produto:
            QMessageBox.warning(self, "Erro", "Produto não encontrado.")
            return

        nome_produto, estoque, valor_unitario, descricao = produto
        estoque = int(estoque)  # Converte para inteiro

        # Substitui vírgulas por pontos no valor_unitario
        valor_unitario = float(valor_unitario.replace(',', '.'))

        # Verifica se a quantidade em estoque é suficiente
        if quantidade_expedida > estoque:
            QMessageBox.warning(self, "Erro", "Quantidade expedida maior que o estoque disponível.")
            return

        # Calcula o valor total
        valor_total = quantidade_expedida * valor_unitario
        
        # Adiciona o item ao carrinho (tabela)
        row_position = self.ui.tableWidget.rowCount()
        self.ui.tableWidget.insertRow(row_position)
        self.ui.tableWidget.setItem(row_position, 0, QTableWidgetItem(str(quantidade_expedida)))
        self.ui.tableWidget.setItem(row_position, 1, QTableWidgetItem(nome_produto))
        self.ui.tableWidget.setItem(row_position, 2, QTableWidgetItem(descricao))
        self.ui.tableWidget.setItem(row_position, 3, QTableWidgetItem(str(estoque)))
        self.ui.tableWidget.setItem(row_position, 4, QTableWidgetItem(str(valor_unitario)))
        self.ui.tableWidget.setItem(row_position, 5, QTableWidgetItem(str(valor_total)))
        
        self.atualizar_total()
    def atualizar_total(self):
        total = 0.0
        for row in range(self.ui.tableWidget.rowCount()):
            valor_total_item = self.ui.tableWidget.item(row, 4).text()
            total += float(valor_total_item)
        
        self.ui.label_24.setText(f"Total: R$ {total:.2f}")
                # Atualiza o total a pagar (label_28) inicialmente com o subtotal
        self.ui.label_28.setText(f"Total: R$ {total:.2f}")
    def excluir_item_selecionado_venda(self):
        # Obter as linhas selecionadas
        linhas_selecionadas = self.ui.tableWidget.selectionModel().selectedRows()
        
        # Verificar se alguma linha foi selecionada
        if linhas_selecionadas:
            # Iterar sobre as linhas selecionadas em ordem reversa para evitar problemas de índice
            for linha in sorted(linhas_selecionadas, reverse=True):
                self.ui.tableWidget.removeRow(linha.row())
                self.atualizar_total()
        else:
            QMessageBox.warning(self, "Atenção", "Nenhuma linha selecionada para excluir.")
    def aplicar_desconto(self):
        desconto = self.ui.lineEdit_12.text()

        # Verifica se o desconto é válido
        if not desconto.replace('.', '', 1).isdigit():
            QMessageBox.warning(self, "Erro", "Desconto inválido.")
            return

        desconto = float(desconto)

        # Extrair o subtotal do label_24, removendo a parte do texto
        subtotal_text = self.ui.label_24.text().replace('Total: R$ ', '')
        subtotal = float(subtotal_text)
        
        total_com_desconto = subtotal - desconto

        # Atualiza o label_25 e label_28 com os novos valores
        self.ui.label_25.setText(f"{desconto:.2f}")
        self.ui.label_28.setText(f"Total: R$ {total_com_desconto:.2f}")
    def registrar_venda(self):
        cliente = self.ui.comboBox_3.currentText()
        forma_pagamento = self.ui.comboBox_2.currentText()
        desconto = self.ui.lineEdit_12.text()
        valor_total = float(self.ui.label_28.text().split("R$")[1].strip())
        valor_pago = self.ui.lineEdit_11.text()
        
        if not desconto.isdigit():
            desconto = 0
        else:
            desconto = float(desconto)

        if not valor_pago.replace('.', '', 1).isdigit():
            QMessageBox.warning(self, "Erro", "Valor pago inválido.")
            return

        valor_pago = float(valor_pago)

        if valor_pago < valor_total:
            QMessageBox.warning(self, "Erro", "Valor pago é inferior ao valor total a pagar.")
            return

        troco = valor_pago - valor_total if valor_pago > valor_total else 0

        data_venda = datetime.now().strftime("%d/%m/%y")

        # Inserir informações de cada item no banco de dados
        for row in range(self.ui.tableWidget.rowCount()):
            qtd = self.ui.tableWidget.item(row, 0).text()
            produto = self.ui.tableWidget.item(row, 1).text()
            descricao = self.ui.tableWidget.item(row, 2).text()
            valor_total_unit = self.ui.tableWidget.item(row, 3).text()
            subtotal_produt = self.ui.tableWidget.item(row, 4).text()
            codigo = self.ui.lineEdit_6.text()
            cod_venda = self.cod_venda
            self.cursor.execute("""
                INSERT INTO registro_venda (qtd, cod_venda, codigo, produto, descricao, cliente, forma_pagamento, data_venda, valor_total_unit, subtotal_produt, desconto, valor_total, valor_pago, troco)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (qtd, cod_venda, codigo, produto, descricao, cliente, forma_pagamento, data_venda, valor_total_unit, subtotal_produt, desconto, valor_total, valor_pago, troco))

        self.db_connection.commit()
        
        if troco > 0:
            QMessageBox.information(self, "Venda registrada com sucesso", f"Venda registrada com sucesso. Troco: R$ {troco:.2f}")
        else:
            QMessageBox.information(self, "Sucesso", "Venda registrada com sucesso.")

        # Limpar tabela e campos após registrar a venda
        self.ui.tableWidget.setRowCount(0)
        self.ui.lineEdit_6.clear()
        self.ui.lineEdit_9.clear()
        self.ui.lineEdit_12.clear()
        self.ui.lineEdit_11.clear()
        self.ui.label_24.setText("Total: R$ 0.00")
        self.ui.label_25.setText("Desconto: R$ 0.00")
        self.ui.label_28.setText("Total a Pagar: R$ 0.00")
    def exportar_excel(self):
        if self.ui.tableWidget_3.rowCount() == 0:
            QMessageBox.warning(self, "Erro", "A tabela está vazia. Não há nada para exportar.")
            return

        path, _ = QFileDialog.getSaveFileName(self, "Salvar Relatório", "", "Excel Files (*.xlsx);;All Files (*)")
        if not path:
            return

        # Criar um DataFrame a partir da QTableWidget
        data = []
        for row in range(self.ui.tableWidget_3.rowCount()):
            rowData = []
            for column in range(self.ui.tableWidget_3.columnCount()):
                item = self.ui.tableWidget_3.item(row, column)
                if item is not None:
                    rowData.append(item.text())
                else:
                    rowData.append('')
            data.append(rowData)

        headers = [self.ui.tableWidget_3.horizontalHeaderItem(i).text() for i in range(self.ui.tableWidget_3.columnCount())]
        df = pd.DataFrame(data, columns=headers)
        
        df.to_excel(path, index=False)
        QMessageBox.information(self, "Sucesso", "Relatório exportado com sucesso.")
        # Função para exibir todos os registros de venda
    def exibir_registros_venda(self):
        self.cursor.execute("""
            SELECT qtd, cod_venda, codigo, produto, descricao, cliente, forma_pagamento, data_venda, Valor_total_unit, subtotal_produt, desconto, valor_total, valor_pago, troco 
            FROM registro_venda
        """)
        registros = self.cursor.fetchall()

        # Definir o número de colunas na tableWidget_2
        self.ui.tableWidget_2.setColumnCount(14)
        self.ui.tableWidget_2.setHorizontalHeaderLabels([
            "Qtd", "Cod_venda", "Código", "Produto", 'Descrição', "Cliente", "Forma de Pagamento", 
            "Data de Venda", "Valor por Unidade", "Subtotal do item", "Desconto", "Valor Total", "Valor Pago", "Troco"
        ])

        self.ui.tableWidget_2.setRowCount(0)
        for row_number, registro in enumerate(registros):
            self.ui.tableWidget_2.insertRow(row_number)
            for column_number, data in enumerate(registro):
                self.ui.tableWidget_2.setItem(row_number, column_number, QTableWidgetItem(str(data)))
    # Função para aplicar filtros nos registros de venda
    def aplicar_filtro_vendas(self):
        data_inicio = self.ui.dateEdit_5.date().toString("dd/MM/yy")
        data_final = self.ui.dateEdit.date().toString("dd/MM/yy")
        produto = self.ui.lineEdit_13.text()
        filtro_ordem = self.ui.comboBox_4.currentText()
        codigo_venda = self.ui.lineEdit_16.text()
        query = """
            SELECT qtd, cod_venda, codigo, produto, descricao, cliente, forma_pagamento, data_venda, Valor_total_unit, subtotal_produt, desconto, valor_total, valor_pago, troco 
            FROM registro_venda
            WHERE data_venda BETWEEN ? AND ?
        """
        params = [data_inicio, data_final]
#erro no parametro, verificar
        if produto:
            query += " AND produto LIKE ?"
            params.append(f"%{produto}%")
        if codigo_venda:
            query += " AND cod_venda LIKE ?"
            params.append(f"%{codigo_venda}%")
        if filtro_ordem == "Código":
            query += " ORDER BY codigo"
        elif filtro_ordem == "Cliente":
            query += " ORDER BY cliente"
        elif filtro_ordem == "Produto":
            query += " ORDER BY produto"
        
        self.cursor.execute(query, params)
        registros = self.cursor.fetchall()
        print(registros)

        self.ui.tableWidget_2.setRowCount(0)
        for row_number, registro in enumerate(registros):
            self.ui.tableWidget_2.insertRow(row_number)
            for column_number, data in enumerate(registro):
                self.ui.tableWidget_2.setItem(row_number, column_number, QTableWidgetItem(str(data)))
    def exportar_vendas_excel(self):
        # Verificar se há dados na tableWidget_2
        if self.ui.tableWidget_2.rowCount() == 0:
            QMessageBox.warning(self, "Erro", "A tabela está vazia. Não há nada para exportar.")
            return

        # Obter o caminho do arquivo para salvar
        path, _ = QFileDialog.getSaveFileName(self, "Salvar Relatório", "", "Excel Files (*.xlsx);;All Files (*)")
        if not path:
            return

        # Criar um DataFrame a partir da QTableWidget
        data = []
        for row in range(self.ui.tableWidget_2.rowCount()):
            rowData = []
            for column in range(self.ui.tableWidget_2.columnCount()):
                item = self.ui.tableWidget_2.item(row, column)
                rowData.append(item.text() if item is not None else '')
            data.append(rowData)

        headers = [self.ui.tableWidget_2.horizontalHeaderItem(i).text() for i in range(self.ui.tableWidget_2.columnCount())]
        df = pd.DataFrame(data, columns=headers)
        
        # Salvar o DataFrame em um arquivo Excel
        df.to_excel(path, index=False)
        QMessageBox.information(self, "Sucesso", "Relatório exportado com sucesso.")
    def gerar_certificado_os(self):
        selected_items = self.ui.tableWidget_2.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Erro", "Nenhum item selecionado na tabela.")
            return

        # Obter o caminho do modelo de certificado
        template_path, _ = QFileDialog.getOpenFileName(self, "Selecionar Modelo de Certificado", "", "Excel Files (*.xlsx);;All Files (*)")
        if not template_path:
            return

        # Obter o caminho para salvar o certificado preenchido
        save_path, _ = QFileDialog.getSaveFileName(self, "Salvar Certificado", "", "Excel Files (*.xlsx);;All Files (*)")
        if not save_path:
            return

        # Carregar o modelo de certificado
        workbook = load_workbook(template_path)
        sheet = workbook.active

        # Selecionar o cliente a partir do primeiro item selecionado
        cliente_item = selected_items[0]
        cliente = self.ui.tableWidget_2.item(cliente_item.row(), 3).text()  # Coluna "cliente"

        # Consultar o banco de dados para obter as informações do cliente
        self.cursor.execute("""
            SELECT endereco, numero, cep, telefone, cnpj_cpf
            FROM registro_cliente
            WHERE nome = ?
        """, (cliente,))
        cliente_data = self.cursor.fetchone()
        endereco = cliente_data[0] if cliente_data else ""
        numero = cliente_data[1] if cliente_data else ""
        cep = cliente_data[2] if cliente_data else ""
        telefone = cliente_data[3] if cliente_data else ""
        cnpj_cpf = cliente_data[4] if cliente_data else ""

        # Preencher as informações do cliente no modelo
        sheet["C6"] = cliente
        sheet["F6"] = endereco
        sheet["G6"] = numero
        sheet["H6"] = cep
        sheet["C8"] = telefone
        sheet["F8"] = cnpj_cpf
        sheet["C11"] = "VENDAS"

        # Função auxiliar para verificar se uma célula está mesclada
        def is_merged_cell(sheet, row, col):
            for merged_range in sheet.merged_cells.ranges:
                if (merged_range.min_row <= row <= merged_range.max_row and
                        merged_range.min_col <= col <= merged_range.max_col):
                    return True
            return False

        # Mapear colunas da tabela para células no Excel
        column_mapping = {
            'C': 0,  # Qtd
            'D': 2,  # Produto
            'E': 4,  # Forma de pagamento
            'F': 6,  # Desconto
            'G': 7,  # Valor total
            'H': 7   # Valor pago
        }

        # Preencher as informações dos itens selecionados
        max_items = 10
        selected_rows = list(set(item.row() for item in selected_items))[:max_items]  # Filtrar linhas únicas

        for i, row in enumerate(selected_rows):
            excel_row = 15 + i
            for excel_col, table_col in column_mapping.items():
                table_item = self.ui.tableWidget_2.item(row, table_col)
                cell_col = openpyxl.utils.column_index_from_string(excel_col)
                if not is_merged_cell(sheet, excel_row, cell_col):
                    sheet.cell(row=excel_row, column=cell_col, value=table_item.text() if table_item else "")

        # Salvar o arquivo preenchido
        workbook.save(save_path)
        QMessageBox.information(self, "Sucesso", "Certificado de OS gerado com sucesso.")
    def carregar_dados_produto(self):
        # Configurar as colunas da tabela
        column_headers = [
            "Id", "Código", "Produto", "Marca", "Data de entrada", "Grupo", "Peso", "Medida", 
            "Fornecedor", "Lote", "Estoque", "Estoque critico", "Valor de venda", "Valor de Compra", "Descrição"
        ]
        self.ui.tableWidget_7.setColumnCount(len(column_headers))
        self.ui.tableWidget_7.setHorizontalHeaderLabels(column_headers)

        # Consulta para obter todos os produtos do banco de dados
        self.cursor.execute("SELECT id, codigo, nome, marca, data_entrada, grupo, peso, medida, fornecedor, lote, estoque, alerta_reposicao, valor_venda, valor_compra, descricao FROM registro_novo_produto")
        produtos = self.cursor.fetchall()

        # Limpar a tabela antes de preencher
        self.ui.tableWidget_7.setRowCount(0)

        # Preencher a tabela com os produtos
        for row_number, produto in enumerate(produtos):
            self.ui.tableWidget_7.insertRow(row_number)
            for column_number, data in enumerate(produto):
                self.ui.tableWidget_7.setItem(row_number, column_number, QTableWidgetItem(str(data)))
    def consultar_dados_produto(self):
        filtro = self.ui.lineEdit.text()

        if not filtro:
            QMessageBox.warning(self, "Erro", "Por favor, insira um valor para a consulta.")
            return

        # Consulta para obter os produtos filtrados do banco de dados
        query = """
        SELECT id, codigo, nome, marca, data_entrada, grupo, peso, medida, fornecedor, lote, estoque, alerta_reposicao, valor_venda, valor_compra, descricao 
        FROM registro_novo_produto 
        WHERE codigo LIKE ? OR nome LIKE ? OR marca LIKE ? OR grupo LIKE ? OR fornecedor LIKE ? OR descricao LIKE ?
        """
        like_filtro = f'%{filtro}%'
        self.cursor.execute(query, (like_filtro, like_filtro, like_filtro, like_filtro, like_filtro, like_filtro))
        produtos = self.cursor.fetchall()

        # Configurar as colunas da tabela
        column_headers = [
            "Id", "Código", "Produto", "Marca", "Data de entrada", "Grupo", "Peso", "Medida", 
            "Fornecedor", "Lote", "Estoque", "Estoque critico", "Valor de venda", "Valor de Compra", "Descrição"
        ]
        self.ui.tableWidget_7.setColumnCount(len(column_headers))
        self.ui.tableWidget_7.setHorizontalHeaderLabels(column_headers)

        # Limpar a tabela antes de preencher
        self.ui.tableWidget_7.setRowCount(0)

        # Preencher a tabela com os produtos filtrados
        for row_number, produto in enumerate(produtos):
            self.ui.tableWidget_7.insertRow(row_number)
            for column_number, data in enumerate(produto):
                self.ui.tableWidget_7.setItem(row_number, column_number, QTableWidgetItem(str(data)))

        if not produtos:
            QMessageBox.information(self, "Informação", "Nenhum produto encontrado com o filtro fornecido.")
    def atualizar_registro_produto(self):
        # Configurar as colunas da tabela
        column_headers = [
            "Id", "Código", "Produto", "Marca", "Data de entrada", "Grupo", "Peso", "Medida", 
            "Fornecedor", "Lote", "Estoque", "Estoque critico", "Valor de venda", "Valor de Compra", "Descrição"
        ]
        self.ui.tabela_entrada.setColumnCount(len(column_headers))
        self.ui.tabela_entrada.setHorizontalHeaderLabels(column_headers)

        # Consulta para obter todos os produtos do banco de dados
        self.cursor.execute("SELECT id, codigo, nome, marca, data_entrada, grupo, peso, medida, fornecedor, lote, estoque, alerta_reposicao, valor_venda, valor_compra, descricao FROM registro_novo_produto")
        produtos = self.cursor.fetchall()

        # Limpar a tabela antes de preencher
        self.ui.tabela_entrada.setRowCount(0)

        # Preencher a tabela com os produtos
        for row_number, produto in enumerate(produtos):
            self.ui.tabela_entrada.insertRow(row_number)
            for column_number, data in enumerate(produto):
                self.ui.tabela_entrada.setItem(row_number, column_number, QTableWidgetItem(str(data)))
    def atualizar_banco_produto(self):
        # Confirmação antes de atualizar
        reply = QMessageBox.question(
            self, "Confirmação", "Você tem certeza que deseja atualizar os registros?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.No:
            return

        # Percorrer todas as linhas da tabela e atualizar o banco de dados
        for row in range(self.ui.tabela_entrada.rowCount()):
            id_ = self.ui.tabela_entrada.item(row, 0).text()
            codigo = self.ui.tabela_entrada.item(row, 1).text()
            nome = self.ui.tabela_entrada.item(row, 2).text()
            marca = self.ui.tabela_entrada.item(row, 3).text()
            data_entrada = self.ui.tabela_entrada.item(row, 4).text()
            grupo = self.ui.tabela_entrada.item(row, 5).text()
            peso = self.ui.tabela_entrada.item(row, 6).text()
            medida = self.ui.tabela_entrada.item(row, 7).text()
            fornecedor = self.ui.tabela_entrada.item(row, 8).text()
            lote = self.ui.tabela_entrada.item(row, 9).text()
            estoque = self.ui.tabela_entrada.item(row, 10).text()
            alerta_reposicao = self.ui.tabela_entrada.item(row, 11).text()
            valor_venda = self.ui.tabela_entrada.item(row, 12).text()
            valor_compra = self.ui.tabela_entrada.item(row, 13).text()
            descricao = self.ui.tabela_entrada.item(row, 14).text()

            # Atualizar o banco de dados
            self.cursor.execute("""
                UPDATE registro_novo_produto
                SET codigo = ?, nome = ?, marca = ?, data_entrada = ?, grupo = ?, peso = ?, medida = ?, 
                    fornecedor = ?, lote = ?, estoque = ?, alerta_reposicao = ?, valor_venda = ?, valor_compra = ?, descricao = ?
                WHERE id = ?
            """, (codigo, nome, marca, data_entrada, grupo, peso, medida, fornecedor, lote, estoque, alerta_reposicao, valor_venda, valor_compra, descricao, id_))

        self.db_connection.commit()
        QMessageBox.information(self, "Sucesso", "Registros atualizados com sucesso.")
    def pesquisar_produto_entrada(self):
        # Obter o valor do comboBox e do lineEdit
        criterio = self.ui.comboBox.currentText()
        valor_pesquisa = self.ui.lineEdit_17.text().strip()

        # Verificar se o campo de pesquisa está vazio
        if not valor_pesquisa:
            QMessageBox.warning(self, "Erro", "Por favor, insira um valor para pesquisa.")
            return

        # Definir a consulta SQL com base no critério selecionado
        if criterio == "Código":
            self.cursor.execute("SELECT * FROM registro_novo_produto WHERE codigo LIKE ?", ('%' + valor_pesquisa + '%',))
        elif criterio == "Produto":
            self.cursor.execute("SELECT * FROM registro_novo_produto WHERE nome LIKE ?", ('%' + valor_pesquisa + '%',))
        else:
            QMessageBox.warning(self, "Erro", "Critério de pesquisa inválido.")
            return

        produtos = self.cursor.fetchall()

        # Limpar a tabela antes de preencher
        self.ui.tabela_entrada.setRowCount(0)

        # Preencher a tabela com os produtos encontrados
        for row_number, produto in enumerate(produtos):
            self.ui.tabela_entrada.insertRow(row_number)
            for column_number, data in enumerate(produto):
                self.ui.tabela_entrada.setItem(row_number, column_number, QTableWidgetItem(str(data)))

        # Exibir uma mensagem caso nenhum produto seja encontrado
        if not produtos:
            QMessageBox.information(self, "Nenhum Resultado", "Nenhum produto encontrado com os critérios fornecidos.")
    def excluir_registro_produto(self):
        # Obter os itens selecionados na tabela
        selected_items = self.ui.tabela_entrada.selectedItems()

        if not selected_items:
            QMessageBox.warning(self, "Erro", "Por favor, selecione um ou mais itens para excluir.")
            return

        # Exibir uma mensagem de confirmação inicial
        confirm = QMessageBox.question(self, "Confirmar Exclusão", "Você tem certeza que deseja excluir os itens selecionados? Esta ação não pode ser desfeita.", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)     
        if confirm == QMessageBox.StandardButton.No:
            return
        # Criar uma lista de IDs para excluir
        ids_to_delete = set()
        for item in selected_items:
            ids_to_delete.add(self.ui.tabela_entrada.item(item.row(), 0).text())
        # Confirmar novamente antes de excluir
        confirm_final = QMessageBox.question(self, "Confirmar Exclusão", f"Você está prestes a excluir {len(ids_to_delete)} itens. Deseja continuar?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if confirm_final == QMessageBox.StandardButton.No:
            return
        # Excluir os itens do banco de dados
        for id in ids_to_delete:
            self.cursor.execute("DELETE FROM registro_novo_produto WHERE id = ?", (id,))
        self.db_connection.commit()
        # Atualizar a tabela
        self.atualizar_registro_produto()
        QMessageBox.information(self, "Sucesso", "Itens excluídos com sucesso.")
    def importar_dados(self):
        # Abrir o diálogo de arquivo para selecionar a planilha
        path, _ = QFileDialog.getOpenFileName(self, "Selecionar Planilha", "", "Excel Files (*.xlsx);;All Files (*)")
        try:
            if not path:
                return
            
            # Ler a planilha usando pandas
            df = pd.read_excel(path)
            
            # Converter datas para strings no formato apropriado
            if 'Data de Entrada' in df.columns:
                df['Data de Entrada'] = pd.to_datetime(df['Data de Entrada'], errors='coerce').dt.strftime('%Y-%m-%d')
            
            # Verificar se as colunas esperadas estão presentes
            expected_columns = ["Código", "Produto", "Marca", "Data de Entrada", "Grupo", "Peso", "Medida", "Fornecedor", "Lote", "Estoque", "Estoque_Mínimo", "Descrição", "Valor de venda", "Valor de compra"]
            for column in expected_columns:
                if column not in df.columns:
                    QMessageBox.warning(self, "Erro", f"A coluna '{column}' não está presente na planilha.")
                    return
            
            # Inserir dados no banco de dados
            for _, row in df.iterrows():
                self.cursor.execute("""
                    INSERT INTO registro_novo_produto (codigo, nome, marca, data_entrada, grupo, peso, medida, fornecedor, lote, estoque, alerta_reposicao, descricao, valor_venda, valor_compra)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    row["Código"], row["Produto"], row["Marca"], row["Data de Entrada"], row["Grupo"], row["Peso"], row["Medida"], 
                    row["Fornecedor"], row["Lote"], row["Estoque"], row["Estoque_Mínimo"], row["Descrição"], 
                    row["Valor de venda"], row["Valor de compra"]
                ))
            self.db_connection.commit()
            QMessageBox.information(self, "Sucesso", "Dados importados com sucesso.")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Falha ao importar dados: {e}")
#########################################################################################################
    def limpar_campos(self):
        # Limpar apenas o texto dos campos QLineEdit de cliente
        self.ui.codigo_cliente.clear()
        self.ui.nome_cliente.clear()
        self.ui.telefone_cliente.clear()
        self.ui.email_cliente.clear()
        self.ui.cnpcpf_cliente.clear()
        self.ui.endereco_cliente.clear()
        self.ui.cep_cliente.clear()
        self.ui.cidade_cliente.clear()
        self.ui.numero_cliente.clear()
        self.ui.complemento_cliente.clear()

        # Limpar apenas o texto dos campos QLineEdit de fornecedor
        self.ui.edit_cod_forn.clear()
        self.ui.edit_forn_nome.clear()
        self.ui.edit_forn_tele.clear()
        self.ui.edit_email_.clear()
        self.ui.edit_forn_cnpjcpf.clear()
        self.ui.prod_edit_forn.clear()
        self.ui.valor_edit_forn.clear()
        self.ui.mais_edit_forn.clear()
        self.ui.ender_edit_forn.clear()
        self.ui.cep_edit_forn.clear()
        self.ui.cid_edit_forn.clear()
        self.ui.num_edit_forn.clear()
        self.ui.comp_edit_forn.clear()

        # Limpar apenas o texto dos campos QLineEdit de produto
        self.ui.codigo.clear()
        self.ui.nome.clear()
        self.ui.marca.clear()
        self.ui.grupo.clear()
        self.ui.peso.clear()
        self.ui.medida.clear()
        self.ui.estoque.clear()
        self.ui.lote_produto.clear()
        self.ui.reposicao.clear()
        self.ui.venda.clear()
        self.ui.lineEdit_2.clear()
        self.ui.descricao.clear()
        #limpar saida de produtos
        self.ui.lineEdit_8.clear()
        self.ui.lineEdit_5.clear()
        self.ui.lineEdit_4.clear()
        self.ui.lineEdit_3.clear()
        self.ui.lineEdit_10.clear()
        self.ui.lineEdit_7.clear()
        self.ui.volume_atual_saida.clear()
        self.ui.volume_minimo_saida.clear()
        self.ui.lote_saida.clear()
        self.ui.peso_saida.clear()
        self.ui.medida_saida.clear()
def gerar_numero_aleatorio(inicio, fim):
    return random.randint(inicio, fim)
if __name__ == "__main__":
    app = QApplication(sys.argv)
    mainWindow = MainWindow()
    mainWindow.show()
    sys.exit(app.exec())
